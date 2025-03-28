"""
по факту - main версия catboost как мне нравится практически,
только много мусора в коде
"""


# === БЛОК 1: Импорты библиотек ===
import optuna
from optuna.visualization import plot_optimization_history, plot_param_importances, plot_contour
import pandas as pd
import numpy as np
from catboost import CatBoostRegressor, Pool
from sklearn.model_selection import train_test_split
from sklearn.metrics import mean_squared_error, mean_absolute_error, r2_score
from sklearn.inspection import permutation_importance
import matplotlib.pyplot as plt
import seaborn as sns
import joblib
import plotly.express as px
import openpyxl
from openpyxl.styles import PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows

# === БЛОК 2: Функции для Optuna ===


def objective(trial, X_train, X_val, y_train, y_val, cat_features):
    params = {
        'iterations': trial.suggest_int('iterations', 5000, 15000),
        'learning_rate': trial.suggest_float('learning_rate', 1e-5, 0.3, log=True),
        'depth': trial.suggest_int('depth', 6, 12),
        'l2_leaf_reg': trial.suggest_float('l2_leaf_reg', 1e-3, 10, log=True),
        'bagging_temperature': trial.suggest_float('bagging_temperature', 0.0, 10.0),
        'random_strength': trial.suggest_float('random_strength', 1e-3, 10),
        'grow_policy': trial.suggest_categorical('grow_policy', ['SymmetricTree', 'Depthwise', 'Lossguide'])
    }

    model = CatBoostRegressor(
        **params,
        cat_features=cat_features,
        verbose=500,
        early_stopping_rounds=1000,
        task_type="GPU",
        devices='0'
    )

    model.fit(X_train, y_train, eval_set=(X_val, y_val), verbose=False)
    y_pred = model.predict(X_val)
    return mean_squared_error(y_val, y_pred)

# === БЛОК 3: Основная функция обучения ===


def process_and_train_model(file_path, test_size=0.1, threshold_percentage=1,
                            model_save_path="catboost_model.cbm", n_trials=1000):

    # Загрузка данных
    data = pd.read_excel(file_path, engine="openpyxl")

    if 'OFR' not in data.columns:
        raise KeyError("Файл не содержит столбца 'Optimized Flow Rate'.")

    X = data.drop(columns=['OFR'])
    y = data['OFR']

    # Предобработка данных
    X = X.dropna()
    y = y[X.index]

    # Разделение данных
    X_train, X_val, y_train, y_val = train_test_split(
        X, y, test_size=test_size, random_state=42)

    # Определение категориальных признаков
    cat_features = [i for i, col in enumerate(
        X.columns) if X[col].dtype == 'object']

    # Оптимизация гиперпараметров
    study = optuna.create_study(
        direction='minimize', sampler=optuna.samplers.TPESampler(seed=42))
    study.optimize(lambda trial: objective(trial, X_train, X_val, y_train, y_val, cat_features),
                   n_trials=n_trials,
                   gc_after_trial=True,
                   n_jobs=1
                   )

    # Сохранение исследования
    joblib.dump(study, 'optuna_study.pkl')

    # Обучение финальной модели
    best_params = study.best_params.copy()
    best_params.update({
        'cat_features': cat_features,
        'verbose': 500,
        'early_stopping_rounds': 1000,
        'task_type': "GPU",
        'devices': '0'
    })

    final_model = CatBoostRegressor(**best_params)
    final_model.fit(X_train, y_train, eval_set=(X_val, y_val), verbose=False)

    # Оценка модели
    y_pred = final_model.predict(X_val)
    mse = mean_squared_error(y_val, y_pred)
    mae = mean_absolute_error(y_val, y_pred)
    r2 = r2_score(y_val, y_pred)

    # Анализ признаков
    perm_importance = permutation_importance(
        final_model, X_val, y_val, n_repeats=10)
    sorted_idx = perm_importance.importances_mean.argsort()[::-1]

    # Веса признаков из модели
    model_importances = final_model.get_feature_importance()

    feature_analysis = pd.DataFrame({
        'Признак': X.columns[sorted_idx],
        'Пермутационная важность': perm_importance.importances_mean[sorted_idx],
        'Вес из модели': model_importances[sorted_idx],
        'Тип': ['Категориальный' if i in cat_features else 'Числовой' for i in sorted_idx]
    })

    # Сохранение результатов
    final_model.save_model(model_save_path)

    # Генерация отчетов
    generate_reports(study, final_model, X_train, X_val,
                     y_val, y_pred, feature_analysis)

    return final_model, X_train, X_val, y_train, y_val, y_pred, cat_features, feature_analysis, study

# === БЛОК 4: Генерация отчетов ===


def generate_reports(study, model, X_train, X_val, y_val, y_pred, feature_analysis):
    # Создаем единый Excel-файл
    with pd.ExcelWriter('full_report.xlsx', engine='openpyxl') as writer:
        # Лист с важностью признаков
        feature_analysis.to_excel(
            writer, sheet_name='Feature_Importance', index=False)

        # Лист с ошибками
        error_matrix = pd.DataFrame({
            'Истинное значение': y_val.values,
            'Предсказанное значение': y_pred,
            'Абсолютная ошибка': np.abs(y_val - y_pred),
            'Процент ошибки': np.abs((y_val - y_pred) / y_val) * 100
        })
        error_matrix.to_excel(writer, sheet_name='Error_Analysis', index=False)

        # Лист с гиперпараметрами
        hyper_params = pd.DataFrame([study.best_params])
        hyper_params.to_excel(
            writer, sheet_name='Hyperparameters', index=False)

        # Лист с матрицей корреляций
        corr_matrix = X_train.corr()
        corr_matrix.to_excel(writer, sheet_name='Correlation_Matrix')

        # Лист с логами оптимизации
        optimization_log = study.trials_dataframe()
        optimization_log.to_excel(
            writer, sheet_name='Optimization_Log', index=False)

        # Применяем форматирование
        workbook = writer.book
        apply_formatting(workbook, feature_analysis)

    # Сохранение графиков
    save_plots(study, X_train, y_val, y_pred)


def apply_formatting(workbook, feature_analysis):
    # Форматирование для Feature_Importance
    ws = workbook['Feature_Importance']

    # Добавляем рекомендации
    threshold = feature_analysis['Пермутационная важность'].quantile(0.25)
    feature_analysis['Рекомендация'] = np.where(
        feature_analysis['Пермутационная важность'] < threshold,
        'Удалить',
        'Сохранить'
    )

    # Записываем рекомендации в столбец F (6-й столбец)
    for row_idx, (_, row) in enumerate(feature_analysis.iterrows(), start=2):
        ws.cell(row=row_idx, column=6).value = row['Рекомендация']

    # Условное форматирование
    red_fill = PatternFill(start_color='FFC7CE',
                           end_color='FFC7CE', fill_type='solid')
    green_fill = PatternFill(start_color='C6EFCE',
                             end_color='C6EFCE', fill_type='solid')

    for row in range(2, len(feature_analysis)+2):
        cell = ws[f'F{row}']
        if cell.value == 'Удалить':
            cell.fill = red_fill
        else:
            cell.fill = green_fill


def save_plots(study, X_train, y_val, y_pred):
    # График сравнения предсказаний
    plt.figure(figsize=(10, 6))
    sns.regplot(x=y_val, y=y_pred, scatter_kws={'alpha': 0.3})
    plt.plot([y_val.min(), y_val.max()], [
             y_val.min(), y_val.max()], 'k--', lw=2)
    plt.title('Сравнение истинных и предсказанных значений')
    plt.savefig('true_vs_predicted.png', dpi=300)
    plt.close()

    # Матрица корреляций
    plt.figure(figsize=(12, 10))
    sns.heatmap(X_train.corr(), annot=True, fmt=".2f", cmap="coolwarm")
    plt.title('Матрица корреляций признаков')
    plt.savefig('correlation_matrix.png', dpi=300)
    plt.close()

    # Визуализации Optuna
    fig = plot_optimization_history(study)
    fig.write_html("optimization_history.html")

    fig = plot_param_importances(study)
    fig.write_html("param_importances.html")


# === БЛОК 5: Запуск приложения ===
if __name__ == "__main__":
    file_path = "Dataset.xlsx"
    model, X_train, X_val, y_train, y_val, y_pred, cat_features, feature_analysis, study = process_and_train_model(
        file_path=file_path,
        test_size=0.1,
        threshold_percentage=1,
        n_trials=1000
    )

    print("Обучение завершено. Основные результаты:")
    print("- full_report.xlsx - полный отчет с листами:")
    print("  * Feature_Importance - веса и важность признаков")
    print("  * Error_Analysis - анализ ошибок предсказаний")
    print("  * Hyperparameters - лучшие гиперпараметры")
    print("  * Correlation_Matrix - матрица корреляций")
    print("  * Optimization_Log - логи оптимизации")
    print("- Визуализации сохранены в PNG и HTML файлы")

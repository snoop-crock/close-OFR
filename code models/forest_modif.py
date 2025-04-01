import joblib
import optuna
import pandas as pd
import numpy as np
from sklearn.model_selection import train_test_split
from sklearn.ensemble import RandomForestRegressor
from sklearn.metrics import mean_squared_error, mean_absolute_error, r2_score
from sklearn.model_selection import cross_val_score
from scipy.stats import spearmanr
import matplotlib.pyplot as plt
import seaborn as sns
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import warnings
import os

# Создаем папку для результатов, если ее нет
os.makedirs('results', exist_ok=True)

# Загрузка данных
df = pd.read_excel('Dataset_Small_q_i.xlsx')  # Замените на ваш файл

# Предположим, что ваш набор данных имеет целевую переменную 'OFR', а остальные - признаки
X = df.drop(columns=['q_i'])
y = df['q_i']

# Разделим на обучающую и тестовую выборки
X_train, X_test, y_train, y_test = train_test_split(
    X, y, test_size=0.1, random_state=42)

# Создаем Excel-файл для логирования
wb = Workbook()
ws_metrics = wb.active
ws_metrics.title = "Trial Metrics"
ws_metrics.append(['Trial', 'MSE', 'RMSE', 'MAE', 'R2', 'Spearman'])

ws_features = wb.create_sheet("Feature Importance")
ws_corr = wb.create_sheet("Correlation Matrix")

# Функция для тренировки модели


def objective(trial):
    # Определяем гиперпараметры для поиска
    param = {
        'n_estimators': trial.suggest_int('n_estimators', 50, 500),
        'max_depth': trial.suggest_int('max_depth', 10, 50),
        'min_samples_split': trial.suggest_int('min_samples_split', 2, 10),
        'min_samples_leaf': trial.suggest_int('min_samples_leaf', 1, 10),
        'max_features': trial.suggest_categorical('max_features', ['sqrt', 'log2', None]),
        'random_state': 42
    }

    # Создание модели с подобранными гиперпараметрами
    model = RandomForestRegressor(**param)

    # Кросс-валидация для оценки
    scores = cross_val_score(model, X_train, y_train,
                             cv=5, scoring='neg_mean_squared_error')
    mse = -scores.mean()

    # Дополнительные метрики
    model.fit(X_train, y_train)
    y_pred = model.predict(X_test)

    rmse = np.sqrt(mse)
    mae = mean_absolute_error(y_test, y_pred)
    r2 = r2_score(y_test, y_pred)
    spearman = spearmanr(y_test, y_pred)[0]

    # Логирование в Excel
    ws_metrics.append([trial.number, mse, rmse, mae, r2, spearman])

    return mse


# Создание объекта для исследования с помощью Optuna
study = optuna.create_study(direction='minimize')  # Минимизируем MSE
study.optimize(objective, n_trials=100)

# Сохраняем историю trials в DataFrame
trials_df = pd.DataFrame(study.trials_dataframe())
trials_history_path = 'results/trials_history.xlsx'
trials_df.to_excel(trials_history_path, index=False)

# Печать лучших параметров
print('Best parameters:', study.best_params)

# Обучаем модель с лучшими параметрами
best_params = study.best_params
best_model = RandomForestRegressor(**best_params)
best_model.fit(X_train, y_train)

# Предсказания на тестовой выборке
y_pred = best_model.predict(X_test)

# Оценка качества модели
mse = mean_squared_error(y_test, y_pred)
rmse = np.sqrt(mse)
mae = mean_absolute_error(y_test, y_pred)
r2 = r2_score(y_test, y_pred)
spearman = spearmanr(y_test, y_pred)[0]

print("\nFinal Model Metrics:")
print(f"MSE: {mse:.4f}")
print(f"RMSE: {rmse:.4f}")
print(f"MAE: {mae:.4f}")
print(f"R²: {r2:.4f}")
print(f"Spearman Correlation: {spearman:.4f}")

# Важность признаков
feature_importance = pd.DataFrame({
    'Feature': X.columns,
    'Importance': best_model.feature_importances_
}).sort_values('Importance', ascending=False)

# Сохраняем важность признаков в Excel
for r in dataframe_to_rows(feature_importance, index=False, header=True):
    ws_features.append(r)

# Корреляционная матрица
corr_matrix = X.corr()
for r in dataframe_to_rows(pd.DataFrame(corr_matrix), index=True, header=True):
    ws_corr.append(r)

# Визуализация важности признаков
plt.figure(figsize=(10, 6))
sns.barplot(x='Importance', y='Feature', data=feature_importance)
plt.title('Feature Importance')
plt.tight_layout()
plt.savefig('results/feature_importance.png')
plt.close()

# Визуализация корреляционной матрицы
plt.figure(figsize=(12, 10))
sns.heatmap(corr_matrix, annot=True, fmt=".2f", cmap='coolwarm')
plt.title('Correlation Matrix')
plt.tight_layout()
plt.savefig('results/correlation_matrix.png')
plt.close()

# График обучения (история trials)
plt.figure(figsize=(10, 6))
plt.plot(trials_df['number'], trials_df['value'], 'b-', label='MSE')
plt.xlabel('Iteration')
plt.ylabel('MSE')
plt.title('Optimization History')
plt.grid()
plt.savefig('results/optimization_history.png')
plt.close()

# Сохранение Excel файла
wb.save('results/model_metrics.xlsx')

# Сохранение модели
joblib.dump(best_model, 'results/random_forest_optuna_model.pkl')

print("\nAll results saved in 'results' directory")
